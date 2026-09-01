"""
GraceEMO Master Database Enrichment Script — Architecture, Tech Stack & Workflow Diagrams
Generates:
1. Architecture_&_Tech_Stack: Complete matrix of all 9 architectural layers, languages, tools, dependencies, and protocols.
2. Pipeline_Workflow_Diagrams: Comprehensive ASCII/Unicode system architecture, voice, vision, LiDAR, kernel, and kinematics flowcharts.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = "docs/GraceEMO_Master_Database.xlsx"

TECH_STACK_DATA = [
    (
        "LAYER-1",
        "Physical Hardware & Actuation",
        "Chassis, Wheels, Pan/Tilt Neck, Dual 90° Arms",
        "Embedded C / MicroPython",
        "FreeRTOS, Arduino Core, ESP-IDF",
        "Serial, I2C, SPI, PWM, USB-UART",
        "UART 115200 bps, CAN 2.0B, Custom Packet Frames",
        "Binary byte packets, 8-bit registers",
        "NVIDIA Jetson Orin Nano / STM32F4 / ESP32-S3",
        "Directly drives physical differential drive DC gearmotors (0.12m wheel radius, 0.42m track width), dual neck servos (yaw ±1.2 rad, pitch -0.5 to +0.6 rad), dual 90° arm servos (0 to 1.57 rad), and reads wheel encoders.",
        "Deterministic low-latency microsecond motor PWM generation, high-current H-bridge drive, and real-time hardware encoder quadrature decoding."
    ),
    (
        "LAYER-2",
        "Low-Level Safety & Reflex Control",
        "Safety Watchdog & Velocity Multiplexer",
        "C++20 (ISO/IEC 14882:2020)",
        "ROS 2 rclcpp, sensor_msgs, geometry_msgs, std_msgs",
        "GCC 13 / Clang 17, CMake 3.28, Colcon",
        "ROS 2 Intra-Process Comm, CycloneDDS / FastDDS",
        "ROS 2 IDL (Twist, LaserScan, JointState, BodyCommand)",
        "NVIDIA Jetson Orin Nano (6-core ARM Cortex-A78AE)",
        "Runs deterministic 50 Hz control loop (safety_servo_node.cpp). Evaluates raw 360° LiDAR ranges; if min_range < 1.0m or E-STOP is active, immediately overrides /cmd_vel to zero. Clamps neck pan/tilt and hand joints within URDF mechanical safety envelopes.",
        "C++20 compiled native execution guarantees zero garbage collection pauses and sub-millisecond execution time, ensuring a hard physical safety reflex barrier before motors."
    ),
    (
        "LAYER-3",
        "Robot Middleware & Node Graph",
        "ROS 2 Jazzy Jalisco Ecosystem",
        "Python 3.10+ / C++20",
        "ROS 2 Jazzy, rclpy, rclcpp, ament_cmake, ament_python",
        "ros2cli, colcon, rosdep, vcs, tf2_ros",
        "DDS (Data Distribution Service) over UDP/IP, Shared Memory",
        "ROS 2 Interfaces (.msg, .srv, .action)",
        "Edge Jetson Orin Nano / Dev Workstation",
        "Provides modular node discovery, topic publish/subscribe graph, synchronous services (/gracemo/recall, /gracemo/ask_question), and parameter servers for brain, control, memory, sensors, and simulation packages.",
        "Standard robotics industry middleware providing modular decoupling, standardized hardware abstraction, and interoperability across heterogeneous languages and processes."
    ),
    (
        "LAYER-4",
        "High-Performance Event Bus & State Ledger",
        "GRaCEmo Rust Kernel Nervous System",
        "Rust 1.80+ (2021 Edition)",
        "Tokio (async runtime), Axum (HTTP/SSE), Serde, Rusqlite, CorsLayer",
        "Cargo, rustc, clippy, cargo-check",
        "Unix Domain Sockets (/tmp/gracemo.sock), HTTP REST, SSE (/events/live)",
        "JSON (Canonical Event envelopes, UUID v4, UTC timestamps)",
        "Linux host / Container daemon (Port 7780)",
        "Acts as central nervous system: maintains in-memory RwLock live state snapshot, commits immutable audit events into SQLite WAL ledger (~/.gracemo/ledger.db), broadcasts live events via SSE to web clients and adapters, and provides REST /snapshot and /dispatch endpoints.",
        "Rust provides guaranteed memory safety, fearless concurrency, zero GC pauses, and sub-millisecond event throughput. SQLite WAL mode ensures persistent black-box auditability."
    ),
    (
        "LAYER-5",
        "Bridge & Adapter Layer",
        "ROS 2 <-> Kernel Bridge Adapter",
        "Python 3.10+",
        "gracemo_sdk, Requests, SSEClient, rclpy",
        "pip, pyproject.toml, hatchling / setuptools",
        "HTTP POST /emit, SSE stream /events/live, ROS 2 pub/sub",
        "Canonical Event JSON, ROS 2 messages",
        "Edge Jetson / Host Container",
        "Subscribes to ROS 2 topics (/odom, /scan, /gracemo/sensors, /gracemo/detections, /gracemo/voice_command), packages them into canonical Event types, and emits to Kernel. Listens to Kernel SSE for ActionRequested (NavigateTo, Stop, Speak, LookAt) and publishes to ROS 2 actuators.",
        "Decouples core ROS 2 robot execution from the high-concurrency Rust kernel event bus; enables language-agnostic adapter development."
    ),
    (
        "LAYER-6",
        "Cognitive AI, Vision & Language Brain",
        "Multi-Modal Reasoning & Behavioral Planner",
        "Python 3.10+ / PyTorch / TensorRT",
        "Google GenAI SDK (Gemini 2.0 Flash), OpenCV, NumPy, SQLite3, Vosk, Whisper",
        "YOLOv11, TensorRT-LLM, Apple Metal / CUDA 12",
        "REST API (Google Generative AI), ROS 2 services",
        "Structured JSON schema, audio WAV/PCM 16kHz, RGB NumPy arrays",
        "Cloud API (Gemini) + Jetson GPU (YOLO TensorRT)",
        "Extracts natural language user intent via Gemini 2.0 Flash, recalls campus destinations dynamically from SQLite memory, manages multi-step mission state machines, tracks visible humans via YOLO, drives pan/tilt gaze servos, and generates spoken speech confirmations.",
        "Combines cloud multi-modal reasoning (Gemini Flash: 200-400ms latency) with local deterministic safety heuristics for zero-downtime offline fallback."
    ),
    (
        "LAYER-7",
        "Physics Simulation & Digital Twin",
        "Campus Physics & Sensor Raycasting Engine",
        "Python 3.10+, C++",
        "Tornado (Web/WebSocket), OpenCV, NumPy, Gazebo Harmonic, SDF/URDF",
        "gz-sim, xacro, robot_state_publisher",
        "Full-Duplex WebSockets (ws://localhost:8888/ws), TCP/IP",
        "Digital twin state JSON payload (50Hz), Base64 JPEG camera buffers",
        "Host Container / Headless Simulation GPU",
        "Simulates 200m x 200m LPU campus world with 13 structures, raycasts 360° LiDAR scans, generates synthetic 5-feed camera views (front, left, right, depth, detections), computes 4-wheel kinematics and battery drain, executes social force crowd physics, and streams digital twin state to web clients at 50 Hz.",
        "Enables full-scale testing of perception, autonomous navigation, and fleet missions without physical hardware or real-world collision risks."
    ),
    (
        "LAYER-8",
        "Command Center UI & Web 3D Studio",
        "Browser-Native Autonomous Robotics Console",
        "HTML5, CSS3, JavaScript (ES6 Modules)",
        "Three.js (r128 WebGL), Canvas 2D API, Google Web Fonts",
        "Chrome / Safari / Edge DevTools, Vite / Live-Server",
        "WebSocket protocol, HTTP/1.1",
        "JSON state telemetry, JPEG data URLs, DOM event models",
        "Client Web Browser (Desktop, Tablet, Mobile)",
        "Provides operators with an enterprise-grade dark-themed command center. Renders real-time 3D campus WebGL world, 4-wheel drive RPM telemetry gauges, live camera grid, LiDAR radar plots, interactive NL mission builder, active mission step checkpoints (✓/▶/○), D-Pad teleop, and E-STOP trigger.",
        "Zero-install universal accessibility across any device or operating system; eliminates dependency on native Linux desktop installations."
    ),
    (
        "LAYER-9",
        "DevOps, Cloud & Production Deployment",
        "Multi-Container Infrastructure & Database",
        "YAML, Bash, Dockerfile, SQL",
        "Docker Engine, Docker Compose v2, PostgreSQL 16 Alpine, Ubuntu 24.04 LTS",
        "git, docker-compose, psql, pgAdmin",
        "Docker Bridge Network, PostgreSQL TCP 5432, Unix Domain Sockets",
        "SQL DDL/DML, YAML configs, Environment variable profiles (.env)",
        "Workstation / On-Premises Server / Cloud VM",
        "Orchestrates complete robotics digital twin runtime in isolated container. Mounts source workspace with symlink-install, configures headless GPU-accelerated Gazebo rendering, binds PostgreSQL 16 database for historical audit logging, and sets up deterministic ROS 2 discovery boundaries.",
        "Guarantees 100% reproducible environments across development, testing, simulation, and physical robot deployment."
    )
]

DIAGRAMS_BLOCKS = [
    (
        "DIAG-01",
        "End-to-End Tiered System Architecture Diagram",
        "Complete 7-tier blueprint showing hardware, safety mux, ROS 2 graph, Rust kernel, AI brain, simulator, and web studio.",
        """+---------------------------------------------------------------------------------------------------------------+
|                                      GRACEEMO TIERED SYSTEM ARCHITECTURE                                      |
+---------------------------------------------------------------------------------------------------------------+

[ TIER 7: WEB COMMAND CENTER & OPERATOR STUDIO ]
  +-----------------------------------------------------------------------------------------------------------+
  |  Three.js 3D WebGL Canvas (200m Campus)  |  4-Wheel Drive RPM HUD  |  Multi-Camera Grid (Front/L/R/Depth) |
  |  Natural Language Mission Dispatcher     |  Active Step Tracker   |  Emergency Stop (E-STOP) Control  |
  +-----------------------------------------------------------------------------------------------------------+
                                          ^                          |
                   50Hz Telemetry & JPEGs |                          | WebSocket Commands (Teleop, Missions, E-Stop)
                                          |                          v
[ TIER 6: PHYSICS SIMULATION & DIGITAL TWIN ENGINE ] (virtual_space_node.py on Port 8888)
  +-----------------------------------------------------------------------------------------------------------+
  |  * 200m x 200m LPU Campus SDF Model (13 Buildings, Roads, Entrances, Restricted Zones)                   |
  |  * Differential Drive Kinematics (v_l, v_r) -> 4-Wheel RPMs (FL, FR, RL, RR) + Dynamic Battery Drain     |
  |  * 360-Degree LiDAR Raycasting (0.05m - 12.0m) + 5 Synthetic Camera Renderers (Base64 JPEG)              |
  |  * Helbing Social Force Crowd Dynamics (0.8 - 1.4 m/s Pedestrians) + Environmental Weather Presets        |
  +-----------------------------------------------------------------------------------------------------------+
               ^                                   |                                   ^
               | /cmd_vel (Safe Twist)             | /odom, /scan, /camera, /sensors   | /gracemo/nav_goal
               |                                   v                                   |
[ TIER 5: SAFETY CONTROL & VELOCITY MUX ] <---> [ TIER 4: ROS 2 JAZZY NODE GRAPH ] <----+
  +-------------------------------------+       +-------------------------------------------------------------+
  |  safety_servo_node.cpp (C++20 @ 50Hz|       |  * planner_node: InspectorState Reconciler & Action Dispatch|
  |  * Evaluates min_range < 1.0m       |       |  * memory_node: SQLite Persistent Episodic & Place Storage  |
  |  * Latching E-STOP: user_estop_=true|       |  * llm_node: Cognitive Cortex with Gemini 2.0 Flash         |
  |  * Clamps Neck Yaw: [-1.2, 1.2] rad |       |  * mission_system_node: Hierarchical HFSM Task Engine       |
  |  * Clamps Neck Pitch: [-0.5, 0.6]rad|       |  * pedestrian_manager_node: Dynamic Crowd & Vehicle Traffic |
  |  * Clamps 90-Deg Hand: [0, 1.57] rad|       |  * scenario_manager_node: Weather, Fog & Sensor Noise Mod   |
  +-------------------------------------+       +-------------------------------------------------------------+
                                                               ^                             |
                                        ROS 2 Actuator Pubs    |                             | ROS 2 Telemetry Subs
                                        (/body_command, /say)  |                             v
[ TIER 3: BI-DIRECTIONAL BRIDGE ADAPTER ] ---------------------+-----------------------------+
  +-----------------------------------------------------------------------------------------------------------+
  |  gracemo_bridge (RobotBridgeAdapter in adapters/robot-bridge/)                                            |
  |  * Forwarder: ROS 2 Telemetry (/odom, /scan, /sensors, /detections) -> HTTP POST http://127.0.0.1:7780/emit |
  |  * Receiver: Live SSE Stream (/events/live) -> ActionRequested -> Dispatches to ROS 2 Actuators          |
  +-----------------------------------------------------------------------------------------------------------+
                                          ^                          |
                                 REST GET |                          | HTTP POST /emit, /dispatch
                                 SSE Live |                          v
[ TIER 2: HIGH-PERFORMANCE EVENT KERNEL & LEDGER ] (gracemo-kernel daemon on Port 7780)
  +-----------------------------------------------------------------------------------------------------------+
  |  * Tokio Asynchronous Actor Runtime (Zero GC Pause, Microsecond Task Scheduling)                         |
  |  * In-Memory State RwLock (Synchronized Real-Time Snapshot of Pose, Battery, Obstacles, Vision, Voice)  |
  |  * Append-Only SQLite WAL Forensic Ledger (~/.gracemo/ledger.db) -> Permanent Forensic Auditability       |
  |  * Tokio Broadcast Channel (cap 2048) -> Multiplexed SSE Live Event Stream (/events/live)                |
  +-----------------------------------------------------------------------------------------------------------+
                                          |
                                          v
[ TIER 1: PHYSICAL HARDWARE / EMBEDDED ACTUATORS ] (Production Deployment)
  +-----------------------------------------------------------------------------------------------------------+
  |  * 4-Wheel Mobile Chassis: Dual DC Drive Motors (0.12m Wheel Radius) + Dual Ball Casters                 |
  |  * Neck Articulator: Dual Digital Bus Servos (Pan ±69°, Tilt -29° to +34°) + RealSense Depth Camera       |
  |  * Expressive Arms: Dual 90-Degree Digital Pitch Servos (0° Rest, 40° Wave, 90° Full Raise)              |
  |  * NVIDIA Jetson Orin Nano (67 TOPS) Edge Compute + USB Noise-Canceling Microphone Array + I2S Speakers  |
  +-----------------------------------------------------------------------------------------------------------+"""
    ),
    (
        "DIAG-02",
        "Voice Command to Autonomous Navigation Pipeline",
        "Step-by-step pipeline from acoustic speech input to physical waypoint arrival.",
        """[ USER SPEECH ] ---> "Go to the library and deliver these documents"
       |
       v
+-------------------------------------------------------------------+
| 1. AUDIO CAPTURE & STT SUBSYSTEM                                  |
|    - Microphone Array captures 16kHz PCM audio                    |
|    - Energy VAD gates acoustic silence (Hard Gate Theta_min=150k) |
|    - Vosk / Whisper Large v3 Turbo transcribes speech to text     |
|    - Publishes VoiceCommand on /gracemo/voice_command             |
+-------------------------------------------------------------------+
       |
       v
+-------------------------------------------------------------------+
| 2. COGNITIVE CORTEX REASONING (llm_node.py)                       |
|    - Input: Transcript + Live InspectorState Context              |
|    - PRIMARY PATH: Gemini 2.0 Flash via Google GenAI SDK          |
|    - System Prompt enforces strict JSON schema:                   |
|      {                                                            |
|        "answer": "Navigating to Central Library Block 37.",       |
|        "intent": "NAVIGATE",                                      |
|        "confidence": 0.96,                                        |
|        "suggested_actions": ["navigate_to:library"]               |
|      }                                                            |
|    - FALLBACK PATH: Deterministic campus safety heuristics        |
+-------------------------------------------------------------------+
       |
       v
+-------------------------------------------------------------------+
| 3. DYNAMIC MEMORY PLACE RESOLUTION (memory_node.py)               |
|    - Matches extracted landmark key against SQLite facts table    |
|    - Resolves: 'library' -> Coordinates: (45.0, 20.0, "Library")  |
|    - Decoupled from static code; dynamically learned places work  |
+-------------------------------------------------------------------+
       |
       v
+-------------------------------------------------------------------+
| 4. AUTONOMY PLANNER DISPATCH (planner_node.py)                    |
|    - Emits BodyCommand(action='navigate_to', target='library')    |
|    - Publishes target coordinates to /gracemo/nav_goal (Sim)      |
|    - Publishes PoseStamped to /goal_pose (Nav2 Action Server)     |
|    - Speaks spoken confirmation via /gracemo/say                  |
|    - Bridge emits Event(ActionRequested) to Rust Kernel Ledger    |
+-------------------------------------------------------------------+
       |
       v
+-------------------------------------------------------------------+
| 5. 50Hz SAFETY SERVO SUPERVISOR (safety_servo_node.cpp)           |
|    - Validates desired Twist against LiDAR range (min_range > 1m) |
|    - Confirms E-STOP is not active (user_estop_ == false)         |
|    - Passes approved /cmd_vel to physical motor controller        |
+-------------------------------------------------------------------+
       |
       v
+-------------------------------------------------------------------+
| 6. CHASSIS MOVEMENT & ARRIVAL CONFIRMATION                        |
|    - 4-Wheel Differential Drive moves robot toward (45.0, 20.0)   |
|    - Distance error < 0.40m triggers arrival state                |
|    - Halts velocities, raises hand to hand_hi (~40° greeting wave)|
|    - Announces: "I have arrived at Central Library (B37)."        |
+-------------------------------------------------------------------+"""
    ),
    (
        "DIAG-03",
        "Vision Perception, Gaze Tracking & Expressive Wave Pipeline",
        "RGB frame capture to YOLO person tracking, neck pan/tilt gaze, and greeting wave.",
        """[ OPTICAL SENSOR ] ---> RGB Camera Frame (640x480 @ 15 FPS) on /camera/image_raw
       |
       v
+-------------------------------------------------------------------+
| 1. OBJECT DETECTION & DEPTH LOCALIZATION                          |
|    - YOLOv11 TensorRT model detects campus entities               |
|    - Labels: 'person', 'vehicle', 'bicycle', 'door', 'chair'      |
|    - Extracts 2D bounding box: (xmin, ymin, xmax, ymax)           |
|    - Correlates with Depth Frame -> Estimates Distance (meters)   |
|    - Publishes Detection msg on /gracemo/detections               |
+-------------------------------------------------------------------+
       |
       v
+-------------------------------------------------------------------+
| 2. BEARING & PERCEPTION VECTOR COMPUTATION                        |
|    - Normalized Center X: cx = (xmin + xmax) / (2 * width)        |
|    - Angular Bearing: person_bearing = (cx - 0.5) * 1.2 radians   |
|    - Distance: person_distance = depth_sensor_meters              |
|    - Reconciled into InspectorState: person_visible = True        |
+-------------------------------------------------------------------+
       |
       v
+-------------------------------------------------------------------+
| 3. ACTIVE NECK GAZE CONTROL                                       |
|    - planner_node calls look_at(person_bearing, 0.15)             |
|    - Clamped in safety_servo_node: yaw in [-1.2, 1.2], pitch in   |
|      [-0.5, 0.6] radians                                          |
|    - Publishes Float64 to /neck_yaw/cmd_pos & /neck_pitch/cmd_pos |
|    - Camera tilts and pans to center person in visual frame       |
+-------------------------------------------------------------------+
       |
       v
+-------------------------------------------------------------------+
| 4. AUTONOMOUS GREETING & GESTURE TRIGGER                          |
|    - Evaluates Proximity: person_distance < 2.5m                  |
|    - Evaluates Cooldown: time_since_last_greeting > 15.0s         |
|    - Halts forward movement (linear_v = 0.0)                      |
|    - Dispatches arm gesture: hand_hi (~40° wave / 0.70 rad)       |
|    - Dispatches spoken greeting to TTS engine:                    |
|      "Hello! I am GraceEMO. Welcome to campus. How may I assist?" |
+-------------------------------------------------------------------+"""
    ),
    (
        "DIAG-04",
        "360° LiDAR Raycasting & 50Hz Safety Cutoff Pipeline",
        "Continuous distance scanning, safety threshold gating, and collision avoidance.",
        """[ LIDAR SENSOR / SIMULATION RAYCASTER ]
  * 360 Beams, Angular Increment 1.0°, Range 0.05m - 12.0m
  * Publishes LaserScan msg on /scan at 10 Hz
       |
       v
+-------------------------------------------------------------------+
| 1. SCAN RANGE FILTERING & MINIMUM DISTANCE EXTRACTION             |
|    - Iterates over all 360 range floats:                          |
|      for (float r : msg->ranges) {                                |
|        if (std::isfinite(r) && r < mn) mn = r;                    |
|      }                                                            |
|    - min_range_ = isfinite(mn) ? mn : 12.0m;                      |
+-------------------------------------------------------------------+
       |
       v
+-------------------------------------------------------------------+
| 2. 50Hz DETERMINISTIC SAFETY WATCHDOG (safety_servo_node.cpp)     |
|    - Evaluates Safety Invariant every 20ms:                       |
|      const bool blocked = user_estop_ || min_range_ < 1.0m;       |
|                                                                   |
|    [ BLOCKED == TRUE ]                     [ BLOCKED == FALSE ]   |
|            |                                        |             |
|            v                                        v             |
|    Override /cmd_vel to 0.0             Forward desired Twist     |
|    Motors lock in physical halt         Normal movement allowed   |
+-------------------------------------------------------------------+
       |
       v
+-------------------------------------------------------------------+
| 3. SIMULTANEOUS NAV2 DYNAMIC REPLANNING                           |
|    - Costmap 2D receives /scan; obstacle layer marks grid cells   |
|    - Inflation layer inflates lethal obstacle perimeter           |
|    - DWB Local Planner detects path obstruction                   |
|    - Evaluates alternative trajectory arcs to steer around object |
|    - Once bypass clears (min_range > 1.0m), motion resumes        |
+-------------------------------------------------------------------+"""
    ),
    (
        "DIAG-05",
        "Rust Kernel EventBus & Append-Only State Ledger Pipeline",
        "Kernel state ingestion, SQLite WAL ledger persistence, and live SSE event broadcasting.",
        """[ ROS 2 TWIN / EXTERNAL SENSORS ]
  * Telemetry (Pose, Battery, Obstacle, Detections, Voice)
       |
       v
+-------------------------------------------------------------------+
| 1. ADAPTER INGESTION (adapters/robot-bridge/)                     |
|    - Encapsulates telemetry in canonical Event envelope:          |
|      { "id": UUID_v4, "timestamp": UTC, "source": "RobotBridge",  |
|        "event_type": { "type": "RobotPosition", "data": {...} } } |
|    - Sends via HTTP POST http://127.0.0.1:7780/emit               |
|      or Unix Domain Socket /tmp/gracemo.sock                      |
+-------------------------------------------------------------------+
       |
       v
+-------------------------------------------------------------------+
| 2. KERNEL PROCESS EVENT ROUTINE (kernel/gracemo-kernel/src/main.rs|
|    - Step 2a: Acquires write lock on live_state RwLock            |
|               Reconciles pose, battery, obstacle, vision, voice   |
|    - Step 2b: Inserts into SQLite ledger (~/.gracemo/ledger.db)   |
|               PRAGMA synchronous = NORMAL, WAL journal mode       |
|               Creates immutable forensic record with primary key  |
|    - Step 2c: Pushes Event into Tokio broadcast channel (cap 2048)|
+-------------------------------------------------------------------+
       |                                   |
       | Broadcast Fan-Out                 | REST Query
       v                                   v
+-----------------------------+     +-------------------------------+
| 3. SSE LIVE EVENT STREAM    |     | 4. SYNCHRONOUS SNAPSHOT API   |
|    GET /events/live         |     |    GET /snapshot              |
|    Real-time streaming push |     |    Returns instantaneous      |
|    Subscribers: Web UI, SDK |     |    atomic world state JSON    |
+-----------------------------+     +-------------------------------+"""
    ),
    (
        "DIAG-06",
        "4-Wheel Kinematics, Battery Drain & Command Center HUD Pipeline",
        "Velocity commands to individual wheel RPM calculation, power drain, and web visualization.",
        """[ TARGET VELOCITIES ] ---> Linear Velocity (v m/s), Angular Velocity (w rad/s)
       |
       v
+-------------------------------------------------------------------+
| 1. DIFFERENTIAL DRIVE FORWARD KINEMATICS                          |
|    - Track Separation: L = 0.42 meters                            |
|    - Drive Wheel Radius: R = 0.12 meters                          |
|    - Left Ground Speed:  v_left  = v - (w * L / 2)                |
|    - Right Ground Speed: v_right = v + (w * L / 2)                |
+-------------------------------------------------------------------+
       |
       v
+-------------------------------------------------------------------+
| 2. INDIVIDUAL 4-WHEEL RPM COMPUTATION & FAULT INJECTION CHECK     |
|    - Wheel RPM = (v_wheel / (2 * pi * R)) * 60.0                  |
|    - FL Wheel RPM: rpm_left  (Status: OK / FAULT)                 |
|    - FR Wheel RPM: rpm_right (Status: OK / FAULT)                 |
|    - RL Wheel RPM: rpm_left  (Status: OK)                         |
|    - RR Wheel RPM: rpm_right (Status: OK)                         |
|    - If fault active (wheel_fl_failure): FL RPM forced to 0.0     |
+-------------------------------------------------------------------+
       |
       v
+-------------------------------------------------------------------+
| 3. DYNAMIC POWER & BATTERY DRAIN SIMULATION                       |
|    - If moving (|v| > 0.01 or |w| > 0.01):                        |
|      battery = max(5.0, battery - (0.0005 + |v| * 0.002))         |
|    - Dynamic discharge curve reflects authentic campus transit    |
+-------------------------------------------------------------------+
       |
       v
+-------------------------------------------------------------------+
| 4. MONOLITHIC 50Hz WEBSOCKET BROADCAST (Port 8888)                |
|    - Payload includes: robot.wheels.{fl, fr, rl, rr}, pose, scan, |
|      5 camera JPEGs, active mission steps, server, network        |
|    - Web app.js receives JSON: updates 4-Wheel HUD gauges,        |
|      animates Three.js 3D wheel rotations, updates mission state  |
+-------------------------------------------------------------------+"""
    )
]

def style_header_cell(cell, fill_hex="001F4E78", font_size=11):
    cell.font = Font(name="Calibri", size=font_size, bold=True, color="00FFFFFF")
    cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="00D9D9D9")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_data_cell(cell, is_even=False, align_left=True, is_mono=False):
    cell.font = Font(name="JetBrains Mono" if is_mono else "Calibri", size=9 if is_mono else 10, bold=False, color="00000000")
    bg_hex = "00F2F5F8" if is_even else "00FFFFFF"
    cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
    cell.alignment = Alignment(horizontal="left" if align_left else "center", vertical="top", wrap_text=True)
    thin = Side(border_style="thin", color="00D9D9D9")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def populate_sheet(wb, title, headers, data, header_fill="001F4E78", col_widths=None, is_diagram=False):
    if title in wb.sheetnames:
        ws = wb[title]
        wb.remove(ws)
    
    ws = wb.create_sheet(title=title)
    ws.views.sheetView[0].showGridLines = True
    
    # Title Banner
    last_col = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"GraceEMO Autonomous Campus Robot — {title.replace('_', ' ')}"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="00FFFFFF")
    title_cell.fill = PatternFill(start_color="00112233", end_color="00112233", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36
    
    # Subtitle Banner
    ws.merge_cells(f"A2:{last_col}2")
    sub_cell = ws.cell(row=2, column=1)
    sub_cell.value = "Comprehensive Multi-Tier Architecture Blueprint, Detailed Technology Stack Specifications & Visual Operational Dataflow Diagrams"
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="00D9D9D9")
    sub_cell.fill = PatternFill(start_color="00112233", end_color="00112233", fill_type="solid")
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # Blank row
    ws.row_dimensions[3].height = 10
    
    # Header Row
    header_row_idx = 4
    ws.row_dimensions[header_row_idx].height = 30
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=header_row_idx, column=col_idx)
        c.value = h
        style_header_cell(c, fill_hex=header_fill, font_size=10)
    
    # Data Rows
    for row_idx, row_data in enumerate(data, start=header_row_idx + 1):
        ws.row_dimensions[row_idx].height = 380 if is_diagram else 85
        is_even = (row_idx % 2 == 0)
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.value = val
            is_mono_col = is_diagram and (col_idx == 4)
            align_l = (col_idx not in [1, 2])
            style_data_cell(c, is_even=is_even, align_left=align_l, is_mono=is_mono_col)
            
    # Set column widths
    if col_widths:
        for col_idx, w in enumerate(col_widths, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = w

def main():
    print(f"Loading {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # 1. Architecture_&_Tech_Stack
    headers_tech = [
        "Layer ID",
        "Architectural Layer",
        "Subsystem / Component",
        "Programming Languages",
        "Frameworks & Engines",
        "Compilers, Tools & SDKs",
        "Protocols & Inter-Process Comm",
        "Data Formats & Message Schemas",
        "Hardware Target & Execution Env",
        "Exact Functional Purpose in GraceEMO",
        "Key Engineering Rationale (Why This Tech Stack?)"
    ]
    widths_tech = [12, 24, 30, 22, 34, 30, 32, 32, 32, 48, 48]
    
    print("Generating Architecture_&_Tech_Stack sheet...")
    populate_sheet(
        wb,
        title="Architecture_&_Tech_Stack",
        headers=headers_tech,
        data=TECH_STACK_DATA,
        header_fill="001F4E78",
        col_widths=widths_tech,
        is_diagram=False
    )
    
    # 2. Pipeline_Workflow_Diagrams
    headers_diag = [
        "Diagram ID",
        "Architecture / Pipeline Name",
        "Diagram Scope & Description",
        "ASCII System Pipeline & Dataflow Diagram"
    ]
    widths_diag = [14, 32, 42, 115]
    
    print("Generating Pipeline_Workflow_Diagrams sheet...")
    populate_sheet(
        wb,
        title="Pipeline_Workflow_Diagrams",
        headers=headers_diag,
        data=DIAGRAMS_BLOCKS,
        header_fill="002A5B84",
        col_widths=widths_diag,
        is_diagram=True
    )
    
    print(f"Saving updated workbook to {EXCEL_PATH}...")
    wb.save(EXCEL_PATH)
    print("✅ Master Database successfully updated with Architecture, Tech Stack & Workflow Diagrams!")

if __name__ == "__main__":
    main()
