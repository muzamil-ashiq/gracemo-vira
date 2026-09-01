"""
GraceEMO Master Database Enrichment Script
Populates:
1. Code_Modules_Deep_Dive: In-depth technical breakdown of every module/file across kernel, adapters, ROS 2, web, sim.
2. End_to_End_System_Workflows: Detailed technical explanations of all end-to-end operational pipelines.
3. Updates DT_Upgrade_Log with latest Phase 6/7 engineering deliverables.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_PATH = "docs/GraceEMO_Master_Database.xlsx"

MODULES_DATA = [
    (
        "MOD-01",
        "Kernel Nervous System",
        "kernel/gracemo-kernel/src/main.rs",
        "Rust, Tokio, Axum, SQLite",
        "Central high-concurrency event bus daemon and append-only state ledger",
        "AppState, process_event(), health_handler(), snapshot_handler(), emit_handler(), dispatch_handler(), sse_handler(), history_handler()",
        "HTTP POST /emit, /dispatch; Unix Domain Socket /tmp/gracemo.sock; REST query /history",
        "HTTP GET /snapshot, /health; SSE stream /events/live; SQLite database ~/.gracemo/ledger.db",
        "Runs an asynchronous Tokio runtime with an in-memory live state RwLock, an SQLite append-only ledger in WAL mode, a Tokio broadcast channel (capacity 2048) broadcasting Event envelopes, and an Axum HTTP/SSE server on port 7780. Implements strict state reconciliation for robot pose, battery, obstacles, vision, and voice.",
        "Heart of the nervous system; bridges ROS 2 adapters, Python SDK clients, and external monitoring dashboards. Falls back gracefully if subscribers disconnect.",
        "Production Active (v0.0.1)"
    ),
    (
        "MOD-02",
        "Kernel Nervous System",
        "kernel/gracemo-types/src/lib.rs",
        "Rust, Serde, UUID, Chrono",
        "Strongly-typed canonical data contracts and event payload definitions",
        "EventSource enum, RobotAction enum, EventType enum, Event struct with UUID v4 and UTC timestamp",
        "Serialized JSON event payloads from any adapter or subsystem",
        "Strongly-typed Rust types used across kernel and types crate",
        "Defines strongly-typed event schemas for physical awareness (RobotPosition, RobotBattery, ObstacleDetected, NavigationArrived, PersonVisible, ObjectDetected, VoiceDetected, ActionRequested). Uses Serde tag/content conventions for deterministic JSON serialization.",
        "Shared library dependency for gracemo-kernel and future Rust adapters; enforces contract consistency.",
        "Production Active (v0.0.1)"
    ),
    (
        "MOD-03",
        "Bridge & Adapter Layer",
        "adapters/robot-bridge/gracemo_bridge/bridge.py",
        "Python 3.10+, rclpy, gracemo_sdk",
        "Bi-directional bridge translating between ROS 2 topic graph and Rust Kernel EventBus",
        "RobotBridgeAdapter, ROSBridgeNode, _run_action_listener(), _dispatch_action(), on_odom(), on_sensors(), on_detection(), on_voice()",
        "ROS 2 topics: /odom, /scan, /gracemo/sensors, /gracemo/detections, /gracemo/voice_command, /gracemo/inspector_state_json; Kernel SSE /events/live",
        "Kernel HTTP /emit; ROS 2 commands: /cmd_vel, /gracemo/say, /gracemo/nav_goal, /gracemo/body_command",
        "Subscribes to ROS 2 robot telemetry topics, converts messages to canonical Event types, and POSTs them to Kernel /emit. Runs a background SSE streaming thread listening for ActionRequested events (NavigateTo, Stop, Speak, LookAt) and publishes them directly to ROS 2 actuators. Sends 5-second heartbeats.",
        "Crucial link between ROS 2 Digital Twin / Gazebo and the Rust Kernel. Degrades to monitoring mode if rclpy is not present on host.",
        "Production Active (v0.0.1)"
    ),
    (
        "MOD-04",
        "Bridge & Adapter Layer",
        "adapters/sdk/gracemo_sdk/client.py",
        "Python 3.10+, Requests, SSEClient",
        "Standardized lightweight Python client for GRaCEmo Kernel interaction",
        "AdapterClient, emit(), get_snapshot(), listen_actions(), ConfigLoader",
        "Adapter configuration YAML files; Kernel HTTP and SSE endpoints",
        "HTTP requests to Kernel /emit, /snapshot, /events/live",
        "Encapsulates HTTP POST /emit with automatic UUID v4 generation, timestamping, and error handling. Provides an SSE generator (listen_actions) yielding ActionRequested events. Provides synchronous state snapshot retrieval (/snapshot).",
        "Used by robot-bridge, brain adapter, and custom user scripts to communicate with the Kernel without ROS 2 dependencies.",
        "Production Active (v0.0.1)"
    ),
    (
        "MOD-05",
        "Bridge & Adapter Layer",
        "adapters/brain/gracemo_brain/reasoner.py",
        "Python 3.10+, OpenAI SDK, Google GenAI SDK",
        "Multi-modal conversational reasoning adapter grounding LLM outputs in sensory state",
        "BrainAdapter, _init_llm_client(), think_and_respond(), start()",
        "Kernel /snapshot (vision, position, battery); user voice queries",
        "ActionRequested(Speak) emitted to Kernel EventBus",
        "Fetches current world state snapshot from Kernel, builds multi-turn dialogue prompt with recent history and grounded sensory context (e.g. detected persons, campus coordinates), queries configured LLM (NVIDIA NIM, Gemini, or Ollama), and emits ActionRequested(Speak) back to the Kernel.",
        "Connects high-level cloud/local AI models to the Kernel EventBus; works in parallel or as fallback to ROS 2 llm_node.",
        "Production Active (v0.0.1)"
    ),
    (
        "MOD-06",
        "Cognitive Autonomy & Planning",
        "graceemo_ws/src/gracemo_brain/gracemo_brain/planner_node.py",
        "Python 3, ROS 2 rclpy, Nav2, SQLite",
        "Central robot behavioral brain: InspectorState reconciler, place recall, dispatcher, joint manager",
        "InspectorState, AutonomyPlannerNode, control_loop(), publish_joints(), navigate_to(), stop(), speak(), look_at(), hand_hi(), hand_up(), hand_down(), on_voice()",
        "/odom, /scan, /gracemo/sensors, /gracemo/detections, /gracemo/voice_command, /gracemo/known_places, /gracemo/robot_pose",
        "/cmd_vel, /gracemo/cmd_vel_desired, /gracemo/say, /gracemo/nav_goal, /goal_pose, /gracemo/body_command, /gracemo/inspector_state, /joint_states, servo /cmd_pos topics",
        "Maintains unified InspectorState (pose, battery, obstacle, person, last_voice). Reconciles dynamic named places recalled from SQLite memory. Translates voice commands into autonomous waypoint goals for Nav2 (/goal_pose) and virtual space pursuit (/gracemo/nav_goal). Executes collision halts and dispatches 4-way joint commands.",
        "Primary autonomy node; coordinates brain, memory, vision, control, and twin simulation.",
        "Production Active (Enhanced)"
    ),
    (
        "MOD-07",
        "Cognitive Autonomy & Planning",
        "graceemo_ws/src/gracemo_brain/gracemo_brain/llm_node.py",
        "Python 3, ROS 2 rclpy, Google GenAI / Gemini Flash",
        "Natural-language cognitive cortex with Gemini 2.0 Flash primary reasoning path",
        "LLMReasoningNode, _call_gemini(), handle_question(), extract_place_heuristic()",
        "ROS 2 Service: /gracemo/ask_question (question text, InspectorState context)",
        "Service Response: answer, intent, confidence, suggested_actions",
        "Acts as primary cognitive cortex by querying Gemini 2.0 Flash with a structured system prompt grounded in InspectorState telemetry. Enforces a strict JSON schema containing spoken answer, categorical intent (NAVIGATE, STOP, GREET, HAND_HI, etc.), confidence score, and suggested actions. Gracefully degrades to deterministic heuristics if network fails.",
        "Invoked asynchronously by planner_node upon receiving unhandled voice commands.",
        "Production Active (Enhanced)"
    ),
    (
        "MOD-08",
        "Simulation & Digital Twin",
        "graceemo_ws/src/gracemo_gazebo/gracemo_gazebo/virtual_space_node.py",
        "Python 3, ROS 2 rclpy, Tornado, OpenCV, NumPy",
        "200m x 200m LPU campus simulation engine, raycasting, kinematics, and Tornado WebSocket server",
        "VirtualSpaceNode, update_physics(), update_lidar(), update_camera(), notify_websockets(), start_web_server(), on_cmd_vel(), _apply_nav_pursuit()",
        "/cmd_vel, /gracemo/nav_goal, /joint_states, scenario, pedestrian, mission, server, network, and fault topics; WebSocket client messages",
        "/odom, /scan, /imu/data, /camera/image_raw, /gracemo/sensors, /gracemo/robot_pose, /gracemo/detections, WebSocket digital_twin_state (port 8888)",
        "Simulates 200m x 200m LPU campus environment with 13 architectural structures. Raycasts 360-degree LiDAR scans, renders synthetic camera frames (front, left, right, depth, detections), runs differential drive kinematic updates with dynamic battery drain and 4-wheel RPM feedback, resolves collisions, and serves the Command Center UI via full-duplex WebSocket.",
        "Core Digital Twin engine; bridges ROS 2 middleware with browser-based operator command center.",
        "Production Active (Enhanced)"
    ),
    (
        "MOD-09",
        "Command Center UI",
        "graceemo_ws/src/gracemo_gazebo/web/index.html",
        "HTML5, CSS3, Google Inter & JetBrains Mono Fonts",
        "Autonomous Campus Robotics Command Center web application structure",
        "Header stats (SIMULATION mode, latency, GPU, battery), Sidebar navigation, 3D Campus World viewport, Camera Grid, Perception & 4-Wheel HUD, Mission Builder, D-Pad, System Health footer",
        "User clicks, slider changes, keyboard shortcuts, text inputs, voice commands",
        "DOM elements for Three.js rendering, canvas camera feeds, telemetry values, WebSocket command triggers",
        "Provides an enterprise-grade dark-themed command center interface. Features prominent SIMULATION badges, real-time 4-wheel drive telemetry cards (FL, FR, RL, RR RPMs and fault indicators), E-STOP button, multi-camera feeds, interactive mission builder, and diagnostic telemetry tabs.",
        "Rendered directly in operator browser via Tornado static server on port 8888.",
        "Production Active (Enhanced)"
    ),
    (
        "MOD-10",
        "Command Center UI",
        "graceemo_ws/src/gracemo_gazebo/web/app.js",
        "JavaScript (ES6 Modules), WebSocket",
        "Frontend client controller managing telemetry updates, user interactions, and WebSocket communication",
        "connectWs(), applyTwinState(), paintAllFeeds(), dispatchNlMission(), initUi(), sendWs()",
        "WebSocket messages: digital_twin_state from virtual_space_node; operator clicks and keyboard inputs",
        "WebSocket actions: teleop, reset_pose, create_mission_nl, mission_control, set_scenario, set_weather, set_crowd, inject_fault, estop, voice",
        "Establishes a resilient full-duplex WebSocket connection to ws://localhost:8888/ws with auto-reconnection. Dispatches digital twin telemetry to 3D canvas and HUD elements. Updates 4-wheel RPM gauges and mission step checkpoints (✓ Completed, ▶ Current, ○ Pending). Dispatches E-STOP and voice commands.",
        "Binds UI elements to digital twin state; orchestrates user controls with backend simulation.",
        "Production Active (Enhanced)"
    ),
    (
        "MOD-11",
        "Command Center UI",
        "graceemo_ws/src/gracemo_gazebo/web/gazebo3d.js",
        "JavaScript (ES6 Modules), Three.js (r128)",
        "3D WebGL campus renderer, robot kinematics visualizer, and dynamic agent scene graph",
        "GraceGazebo singleton, init(), animate(), setRobot(), setBuildings(), setDynamicAgents(), setScan(), setWeather(), lookAtRos(), focusRobot(), resetView()",
        "Robot pose and joint angles, LiDAR point ranges, dynamic pedestrian/vehicle agent coordinates, weather state",
        "Interactive 3D WebGL scene rendered to canvas #gazeboCanvas",
        "Initializes Three.js perspective camera, directional and ambient lights, sky dome, grid ground, and building bounding geometries based on campus_metadata.json. Renders GraceEMO 3D robot model with articulating neck and arm joints, animated wheels, 360-degree LiDAR ray points, dynamic pedestrian/vehicle markers, and weather particle systems.",
        "Visualizes digital twin state in real time; provides intuitive orbit, pan, and zoom camera controls.",
        "Production Active"
    ),
    (
        "MOD-12",
        "Safety & Actuation Control",
        "graceemo_ws/src/gracemo_control/src/safety_servo_node.cpp",
        "C++20, ROS 2 rclcpp, sensor_msgs, std_msgs",
        "Deterministic low-level safety watchdog, velocity multiplexer, and joint position servo controller",
        "SafetyServoNode, onScan(), onDesired(), onBody(), tick()",
        "/scan (LaserScan), /gracemo/cmd_vel_desired (Twist), /gracemo/body_command (BodyCommand)",
        "/cmd_vel (safe Twist), /joint_states (JointState), /neck_yaw/cmd_pos, /neck_pitch/cmd_pos, /left_hand/cmd_pos, /right_hand/cmd_pos (Float64)",
        "Executes a deterministic 50 Hz control loop (20ms interval). Evaluates real-time LiDAR minimum distance against a hard collision threshold (1.0m) and operator E-STOP state. Overrides desired velocity commands to zero if an obstacle is detected. Clamps neck pan (±1.2 rad), neck pitch (-0.5 to +0.6 rad), and 90° hand pitch (0 to 1.5708 rad).",
        "Final supervisory safety layer before motor actuators; guarantees AI proposals cannot bypass physical constraints.",
        "Production Active (Enhanced)"
    ),
    (
        "MOD-13",
        "Cognitive Memory & Knowledge",
        "graceemo_ws/src/gracemo_memory/gracemo_memory/memory_node.py",
        "Python 3, ROS 2 rclpy, SQLite",
        "Persistent episodic and semantic memory engine with dynamic place recall publishing",
        "MemoryNode, init_database(), seed_places(), get_all_places_dict(), publish_known_places(), handle_remember(), handle_recall(), on_detection(), on_voice()",
        "/gracemo/detections, /gracemo/voice_command, ROS 2 Services: /gracemo/remember, /gracemo/recall",
        "/gracemo/known_places (String JSON), SQLite database /tmp/graceemo_memory.db",
        "Maintains an SQLite database with facts and events tables. Seeds initial campus landmarks and dynamically learns new ones via /gracemo/remember. Publishes all known destination coordinates on /gracemo/known_places every 3 seconds and on updates. Automatically logs human encounters and voice interactions as episodic events.",
        "Provides long-term knowledge retention and enables dynamic semantic navigation for planner_node.",
        "Production Active (Enhanced)"
    ),
    (
        "MOD-14",
        "Robot Modeling & Kinematics",
        "graceemo_ws/src/gracemo_description/urdf/gracemo.urdf.xacro",
        "URDF, Xacro, XML, Gazebo Sim Plugins",
        "Parametric kinematic and physical robot description with sensors and joint controllers",
        "Links: base_footprint, base_link, 4 wheels, 2 casters, hips, torso, neck_yaw, head, camera, imu, hands; Joints: drive wheels, neck_yaw, neck_pitch, left_hand, right_hand",
        "Xacro properties: chassis dimensions, wheel separation (0.42m), wheel radius (0.12m), joint limits",
        "Robot description topic /robot_description, TF2 coordinate transforms, Gazebo physics simulation models",
        "Models GraceEMO-01 physical architecture: 4.5-ft height, white composite shells, black joint actuators, 4-wheel mobile base (2 continuous drive wheels + 2 caster supports). Defines revolute joint limits for neck yaw (±1.2 rad), neck pitch (-0.5 to +0.6 rad), and dual 90° hand joints (0 to 1.5708 rad). Configures Gazebo DiffDrive, JointStatePublisher, and JointPositionController plugins.",
        "Defines the single source of kinematic truth used across ROS 2 TF tree, Gazebo Harmonic, and WebGL visualization.",
        "Production Active (Verified)"
    ),
    (
        "MOD-15",
        "Autonomous Navigation",
        "graceemo_ws/src/gracemo_navigation/config/nav2_params.yaml",
        "YAML, Nav2, ROS 2",
        "Nav2 autonomous navigation, costmap, and path planning parameter specifications",
        "amcl, bt_navigator, controller_server, planner_server, recoveries_server, global_costmap, local_costmap",
        "LiDAR /scan, odometry /odom, map /map, goal pose /goal_pose",
        "Velocity commands /cmd_vel, path visualization /plan, global and local costmap grids",
        "Configures Nav2 navigation stack for campus mobility: DWB Local Planner for obstacle avoidance and dynamic trajectory evaluation, NavFn global planner for Dijkstra/A* path generation, inflation and obstacle layers for 2D costmaps, AMCL particle filter localization, and recovery behaviors (spin, backup, wait).",
        "Directs robot physical movement during autonomous patrol and delivery missions in Gazebo Harmonic.",
        "Production Active"
    ),
    (
        "MOD-16",
        "Environment & Scenarios",
        "graceemo_ws/src/gracemo_scenarios/gracemo_scenarios/scenario_manager_node.py",
        "Python 3, ROS 2 rclpy",
        "Campus environmental condition, weather simulation, and scenario preset manager",
        "ScenarioManagerNode, apply_scenario(), set_weather(), set_crowd(), on_set_scenario(), on_set_weather()",
        "/gracemo/set_scenario, /gracemo/set_weather, /gracemo/set_crowd",
        "/gracemo/scenario_state (String JSON), /gracemo/weather, /gracemo/crowd_density",
        "Manages 10 simulation scenarios (Normal Campus, Crowded Campus, Night Campus, Emergency, etc.) and weather presets (Clear Day, Rainy, Heavy Fog, Dusk, Night). Computes sensor noise factors and LiDAR attenuation based on precipitation/fog and broadcasts updates to pedestrians and twin simulators.",
        "Enables testing of perception, navigation, and human-robot interaction under diverse environmental conditions.",
        "Production Active"
    ),
    (
        "MOD-17",
        "Environment & Scenarios",
        "graceemo_ws/src/gracemo_pedestrians/gracemo_pedestrians/pedestrian_manager_node.py",
        "Python 3, ROS 2 rclpy, NumPy",
        "Dynamic multi-agent pedestrian and vehicle simulation implementing social force behavior",
        "PedestrianManagerNode, update_agents(), apply_social_forces(), spawn_pedestrian(), spawn_vehicle()",
        "/gracemo/scenario_state, /gracemo/crowd_density, /gracemo/robot_pose",
        "/gracemo/dynamic_agents (String JSON)",
        "Simulates realistic dynamic crowds and traffic across campus roads and walkways. Implements Helbing social force physics: agents navigate toward landmark waypoints while exerting repulsive forces against buildings, obstacles, the GraceEMO robot, and neighboring pedestrians. Speeds vary from 0.8 to 1.4 m/s for pedestrians and 3.0 to 6.0 m/s for campus e-carts.",
        "Provides dynamic moving obstacles for robot perception (YOLO) and local costmap obstacle avoidance.",
        "Production Active"
    ),
    (
        "MOD-18",
        "Mission Planning & Execution",
        "graceemo_ws/src/gracemo_missions/gracemo_missions/mission_system_node.py",
        "Python 3, ROS 2 rclpy",
        "Hierarchical mission engine, natural-language mission parser, and state machine executor",
        "MissionSystemNode, parse_natural_language(), create_mission(), update_mission_lifecycle(), on_nl_command(), on_mission_control()",
        "/gracemo/create_mission_nl, /gracemo/mission_control, /gracemo/robot_pose",
        "/gracemo/mission_state (String JSON), /gracemo/nav_goal",
        "Deconstructs natural-language task requests (e.g. 'Deliver mail to Central Library Block 37') into sequential checkpoint steps (Navigate -> Perception Scan -> Interaction/Handoff -> Confirm Arrival). Tracks active progress percentage and step completion, and manages mission pause, resume, abort, and return-to-base states.",
        "Translates high-level operator intent into discrete navigation goals and arm/speech interactions.",
        "Production Active"
    ),
    (
        "MOD-19",
        "Cloud & Fleet Infrastructure",
        "graceemo_ws/src/gracemo_server/gracemo_server/central_server_node.py",
        "Python 3, ROS 2 rclpy, FastAPI / Tornado",
        "Central AI GPU server simulator and multi-robot fleet management orchestrator",
        "CentralServerNode, simulate_server_load(), allocate_fleet_task(), handle_vlm_offload()",
        "/gracemo/robot_state, /gracemo/set_server",
        "/gracemo/server_state (String JSON), /gracemo/fleet_state, /gracemo/digital_twin_state (port 8090)",
        "Simulates a centralized AI GPU cluster (FastAPI/Tornado on port 8090). Emulates multi-robot fleet telemetry aggregation, heavy AI inference queueing (VLM scene understanding, global path optimization), GPU/CPU/VRAM utilization metrics, and server availability states.",
        "Models edge-cloud hybrid robotics architecture; supports fleet deployment across university campuses.",
        "Production Active"
    ),
    (
        "MOD-20",
        "Cloud & Fleet Infrastructure",
        "graceemo_ws/src/gracemo_network_sim/gracemo_network_sim/network_sim_node.py",
        "Python 3, ROS 2 rclpy, NumPy",
        "Wireless network condition emulator and multi-tier edge/cloud failover controller",
        "NetworkSimNode, simulate_packet_dynamics(), evaluate_failover_state(), on_set_network()",
        "/gracemo/set_network",
        "/gracemo/network_state (String JSON), /gracemo/server_status",
        "Simulates real-world wireless degradation: latency (0 to 500ms), jitter (Gaussian distribution), and packet loss (0 to 25%). Automatically manages 3 failover tiers: ONLINE (full server offloading), PARTIAL (cloud AI disabled, local navigation active), and OFFLINE (complete autonomous edge fallback).",
        "Validates robot operational resilience when crossing Wi-Fi dead zones across the 200m campus.",
        "Production Active"
    ),
    (
        "MOD-21",
        "Testing & Resilience Engineering",
        "graceemo_ws/src/gracemo_fault_injection/gracemo_fault_injection/fault_injection_node.py",
        "Python 3, ROS 2 rclpy",
        "Automated robotics fault injection testing engine supporting 15 hardware and sensor failure modes",
        "FaultInjectionNode, inject_fault(), clear_fault(), evaluate_robot_resilience(), on_inject_fault(), on_clear_fault()",
        "/gracemo/inject_fault, /gracemo/clear_fault",
        "/gracemo/fault_state (String JSON), /gracemo/fault_events",
        "Allows operators to dynamically trigger 15 specific subsystem failures: front LiDAR cutoff, camera occlusion, wheel motor stall (FL/FR), high battery drain, localization jump, network loss, and emergency stop. Emits diagnostic events and verifies whether safety supervisors detect and mitigate the fault.",
        "Essential tool for empirical scientific robotics research, reliability benchmarking, and safety verification.",
        "Production Active"
    ),
    (
        "MOD-22",
        "Testing & Resilience Engineering",
        "graceemo_ws/src/gracemo_analytics/gracemo_analytics/analytics_node.py",
        "Python 3, ROS 2 rclpy",
        "Telemetry analytics engine, distance tracker, and structured JSONL replay logger",
        "AnalyticsNode, compute_performance_metrics(), log_structured_event(), export_replay_session()",
        "/odom, /gracemo/mission_state, /gracemo/fault_state, /gracemo/structured_log",
        "/gracemo/analytics (String JSON), log files in /workspace/gracemo_data/logs/",
        "Computes real-time mission success rates, total distance traveled (odometry integration), collision count, replan frequency, and AI detection accuracy. Serializes structured audit logs to JSONL files on disk and supports session recording for flight-recorder playback.",
        "Provides complete transparency and quantitative metrics for robotics system optimization.",
        "Production Active"
    ),
    (
        "MOD-23",
        "Testing & Resilience Engineering",
        "graceemo_ws/src/gracemo_research/gracemo_research/research_node.py",
        "Python 3, ROS 2 rclpy",
        "Scientific robotics research experiment manager with 5 automated empirical benchmark suites",
        "ResearchNode, start_experiment(), evaluate_trial_hypotheses(), export_empirical_dataset()",
        "/gracemo/analytics, /gracemo/research_control",
        "/gracemo/research_state (String JSON)",
        "Automates academic experimental trials across 5 benchmark domains: Cloud vs. Edge Latency, Pedestrian Avoidance Success, Fault Recovery Latency, Weather Degradation, and Energy Efficiency. Executes parameter sweeps and exports CSV/JSON trial results for scientific publications.",
        "Enables standardized peer-reviewed empirical validation of GraceEMO robotics capabilities.",
        "Production Active"
    ),
    (
        "MOD-24",
        "Master System Orchestration",
        "graceemo_ws/src/gracemo_bringup/launch/lpu_digital_twin.launch.py",
        "Python 3, ROS 2 Launch Framework",
        "Master digital twin launch script orchestrating all 11 core simulation and intelligence nodes",
        "generate_launch_description(), add_node() helper",
        "Launch configuration arguments: direct_cmd_vel, actuate_joints, web_port",
        "Executes 11 ROS 2 nodes simultaneously with parameters and shared configs",
        "Spawns the entire GraceEMO LPU Digital Twin platform in a single command: Virtual Space Studio (port 8888), Scenario Manager, Pedestrian Manager, Mission System, Central AI Server (port 8090), Network Sim, Fault Injection, Analytics, Research Framework, Cortex Brain (llm_node + planner_node), and Memory Engine (memory_node).",
        "Single-command turnkey launch for development, simulation demonstrations, and research testing.",
        "Production Active (Enhanced)"
    ),
    (
        "MOD-25",
        "Master System Orchestration",
        "docker/docker-compose.yml",
        "YAML, Docker, Docker Compose",
        "Enterprise containerization defining multi-container robotics digital twin stack",
        "services: graceemo (ROS 2 Jazzy, Gazebo Harmonic, Python), postgres (PostgreSQL 16 Alpine)",
        "Environment variables (.env): GEMINI_API_KEY, OPENAI_API_KEY, ROS_DOMAIN_ID; volume mounts",
        "Container network: port 8888 (Web Studio), 8090 (Central AI), 5432 (PostgreSQL DB)",
        "Orchestrates complete digital twin runtime in isolated container. Mounts source workspace with symlink-install, configures headless GPU-accelerated Gazebo rendering, binds PostgreSQL 16 database for historical audit logging, and sets up deterministic ROS 2 discovery boundaries.",
        "Guarantees reproducible deployment across macOS (Apple Silicon), Ubuntu workstations, and cloud instances.",
        "Production Active"
    )
]

WORKFLOWS_DATA = [
    (
        "WF-01",
        "Natural Language Voice Navigation",
        "Human voice command -> Cognitive Intent -> Landmark Recall -> Waypoint Dispatch -> Kinematics -> Spoken Arrival",
        "User speaks: 'Go to the door' / 'Take me to the library'",
        "Robot navigates smoothly to destination coordinates, halts safely, waves, and confirms arrival via speech",
        "1. Microphone captures user acoustic speech waveform.\n"
        "2. Voice subsystem (or web UI) transcribes text to VoiceCommand msg on /gracemo/voice_command.\n"
        "3. planner_node.on_voice() receives transcript, evaluates intent (NAVIGATE).\n"
        "4. If unknown place, calls llm_node (Gemini 2.0 Flash) on /gracemo/ask_question for structured JSON intent parsing.\n"
        "5. Destination place key is matched dynamically against self.known_places (synced from memory_node SQLite).\n"
        "6. Landmark coordinates (x, y, name) are resolved; planner calls _begin_nav().\n"
        "7. Emits BodyCommand(action='navigate_to') to safety_servo_node.\n"
        "8. Publishes goal to /gracemo/nav_goal (for virtual space pursuit) and /goal_pose (for Nav2 global planner).\n"
        "9. Robot drives on 4 wheels; control loop monitors distance < 0.4m; upon arrival halts and speaks confirmation."
    ),
    (
        "WF-02",
        "Real-Time Human Detection & Expressive Interaction",
        "Camera RGB feed -> YOLOv11 TensorRT -> Person Bearing/Distance -> Neck Gaze Servo -> Welcome Greeting & Wave",
        "Pedestrian enters robot's camera field of view within 2.5 meters",
        "Robot centers its gaze on the pedestrian, executes greeting wave, and speaks welcome message",
        "1. Virtual space (or physical camera) captures 640x480 RGB frame at 15 FPS on /camera/image_raw.\n"
        "2. YOLO perception node runs object detection and publishes Detection msg on /gracemo/detections.\n"
        "3. Detection includes label='person', confidence score, bounding box center_x, and estimated depth distance.\n"
        "4. planner_node.on_detection() calculates angular bearing: person_bearing = (cx - 0.5) * 1.2 rad.\n"
        "5. Updates InspectorState with person_visible=True, person_distance, and person_bearing.\n"
        "6. Calls look_at(person_bearing, 0.15) -> dispatches Float64 to /neck_yaw/cmd_pos to turn head toward person.\n"
        "7. If distance < 2.5m and cooldown elapsed (>15s), calls greet(): halts velocity, raises right arm to hand_hi (~40°), and speaks: 'Hello! I am GraceEMO. Welcome to campus. How may I assist you?'."
    ),
    (
        "WF-03",
        "LiDAR Raycasting & Safety Reflex Override",
        "360° LiDAR Scan -> Minimum Range Evaluation -> Safety Mux Cutoff -> Path Replanning -> Obstacle Clear",
        "Obstacle or dynamic pedestrian moves directly into robot's path (<1.0m)",
        "Robot instantly halts physical drive motors, updates HUD obstacle indicator, and waits or replans",
        "1. virtual_space_node raycasts 360 laser scan beams against campus walls, buildings, and dynamic agents at 10 Hz.\n"
        "2. LaserScan published on /scan; proximity sensor publishes on /gracemo/proximity_front.\n"
        "3. safety_servo_node (C++20 running at 50 Hz) evaluates all finite scan ranges.\n"
        "4. If minimum range < 1.0m or user E-STOP is active, safety mux immediately overrides /cmd_vel to zero.\n"
        "5. planner_node detects state.obstacle_ahead=True, aborts active kinematic pursuit, and announces: 'Obstacle ahead. Stopping.'.\n"
        "6. In Nav2 path planning mode, costmap inflation layer inflates the obstacle, forcing DWB local planner to recompute a clean bypass trajectory.\n"
        "7. Once path clears (min range > 1.0m), safety mux releases velocity lock."
    ),
    (
        "WF-04",
        "Rust Kernel EventBus & Append-Only State Ledger",
        "Telemetry / Events -> Unix Socket / REST -> SQLite WAL Ledger -> Broadcast Channel -> SSE Stream -> Adapters",
        "Any subsystem emits a physical telemetry update or requests an actuation action",
        "Event is permanently committed to SQLite audit ledger and broadcast to all live SSE streaming subscribers in <1ms",
        "1. Source adapter (e.g. robot-bridge or vision) constructs canonical Event envelope with UUID v4 and timestamp.\n"
        "2. Transmits event via Unix domain socket (/tmp/gracemo.sock) or HTTP POST to http://127.0.0.1:7780/emit.\n"
        "3. gracemo-kernel process_event() acquires write lock on live_state RwLock and updates in-memory snapshot.\n"
        "4. Inserts event record (id, timestamp, source, observed_by, event_type, payload) into SQLite ledger (~/.gracemo/ledger.db) with PRAGMA synchronous = NORMAL and WAL journal mode.\n"
        "5. Pushes event into Tokio broadcast channel (capacity 2048).\n"
        "6. Active SSE streaming clients listening on /events/live immediately receive the serialized event envelope.\n"
        "7. REST clients querying /snapshot or /history receive instantaneous synchronized state."
    ),
    (
        "WF-05",
        "4-Wheel Differential Kinematics & Stability Physics",
        "Target Velocities (v, w) -> Left/Right Wheel Speeds -> RPM Computation -> Battery Drain -> Command Center HUD",
        "Robot accelerates, turns, or navigates campus terrain",
        "Command Center HUD displays real-time RPM gauges for all 4 wheels with dynamic power monitoring",
        "1. Kinematic controller receives linear velocity (v) and angular velocity (w).\n"
        "2. Applies differential drive kinematics: v_left = v - (w * L / 2), v_right = v + (w * L / 2), where L = 0.42m wheel separation.\n"
        "3. Wheel rotational speeds are computed: RPM = (v_wheel / (2 * pi * R)) * 60, where R = 0.12m wheel radius.\n"
        "4. FL and RL wheels track Left RPM; FR and RR wheels track Right RPM.\n"
        "5. Checks fault_state: if a simulated fault is injected (e.g. wheel_fl_failure), that wheel's RPM drops to 0 and status switches to 'FAULT'.\n"
        "6. Power model dynamically consumes battery: battery -= (0.0005 + |v| * 0.002) per step.\n"
        "7. State is packed into digital_twin_state.robot.wheels and broadcast over WebSocket to app.js, updating HUD gauges in real time."
    ),
    (
        "WF-06",
        "Operator Command Center & Digital Twin Synchronization",
        "Tornado WebSocket Server -> Browser Client -> Three.js Canvas -> Real-Time Sensor Feeds -> Multi-Cam Display",
        "Operator opens http://localhost:8888 in web browser",
        "Full 3D digital twin of LPU campus, multi-camera feeds, sensor readouts, and mission controls update at 50 Hz",
        "1. Operator navigates to http://localhost:8888; Tornado server serves index.html, style.css, and app.js.\n"
        "2. app.js establishes WebSocket handshake to ws://localhost:8888/ws.\n"
        "3. virtual_space_node registers connection in active_websockets set.\n"
        "4. Server packs complete digital twin state snapshot (robot pose, joints, 360 scan, camera JPEG, depth JPEG, detections JPEG, weather, mission, server, network, faults, logs) into JSON payload at 50 Hz.\n"
        "5. app.js receives payload: updates Three.js scene graph in gazebo3d.js (robot position, wheel rotations, head pan/tilt, pedestrian markers).\n"
        "6. Paints front, left, right, depth, and detection camera frames onto corresponding canvas elements.\n"
        "7. Updates sensor telemetry readouts (X, Y, YAW, Speed, Proximity, Bumper, Battery, 4-Wheel RPMs)."
    ),
    (
        "WF-07",
        "Hierarchical Mission Planning & Checkpoint Execution",
        "Natural Language Request -> Step Decomposition -> State Machine Execution -> Checkpoint Validation -> Completion",
        "Operator enters: 'Deliver mail to Central Library Block 37' in Command Center input",
        "Robot decomposes task into checkpoints, navigates to Block 37, performs mail handoff, and confirms completion",
        "1. Operator inputs natural-language text in Command Center textarea and clicks Dispatch.\n"
        "2. WebSocket sends { action: 'create_mission_nl', text: 'Deliver mail to Central Library Block 37' } to virtual_space_node.\n"
        "3. Published to /gracemo/create_mission_nl; mission_system_node parses archetype: 'delivery'.\n"
        "4. Generates multi-step mission plan:\n"
        "   - Step 1: Navigate to Central Library (B37)\n"
        "   - Step 2: Scan for recipient / receptionist\n"
        "   - Step 3: Announce delivery and perform handoff gesture (hand_hi / hand_up)\n"
        "   - Step 4: Confirm delivery and return to base\n"
        "5. Publishes active mission state on /gracemo/mission_state and dispatches first waypoint to /gracemo/nav_goal.\n"
        "6. Command Center renders active steps with interactive status icons (✓ Completed, ▶ Current, ○ Pending) and animated progress bar.\n"
        "7. Upon checkpoint arrivals, advances state machine until mission completes."
    ),
    (
        "WF-08",
        "Dual-Tier Emergency Stop (E-STOP) Architecture",
        "Operator Clicks E-STOP -> WebSocket Broadcast -> Motion Nullification -> Safety Servo Watchdog -> Audio Alert",
        "Operator clicks red E-STOP button or hardware safety trip occurs",
        "All linear and angular velocities immediately drop to 0, missions abort, and system enters locked emergency state",
        "1. Operator clicks 'E-STOP' button in Command Center UI or presses emergency key.\n"
        "2. app.js immediately transmits { action: 'estop' } via WebSocket and sets local UI badge to 'E-STOP TRIGGERED'.\n"
        "3. virtual_space_node receives estop action: sets status='EMERGENCY_STOP', target_linear_v=0, target_angular_w=0, linear_v=0, angular_w=0, and clears nav_goal.\n"
        "4. Publishes zero Twist on /cmd_vel, publishes 'abort' on /gracemo/mission_control, and publishes 'stop' on /gracemo/speech_input.\n"
        "5. safety_servo_node receives stop BodyCommand: sets user_estop_ = true, latching all velocity outputs to zero.\n"
        "6. planner_node halts active navigation loop, resets current_task to 'IDLE', and speaks: 'Emergency stop activated.'.\n"
        "7. Robot cannot move until operator explicitly clicks 'Clear E-STOP' to verify safety clearance."
    ),
    (
        "WF-09",
        "Multi-Tier Wireless Failover & Degraded Operation",
        "Network Quality Measurement -> Latency/Loss Threshold Evaluation -> Failover State Transition -> Local AI Fallback",
        "Robot moves into campus Wi-Fi dead zone with high packet loss (>15%) or disconnected server",
        "Robot automatically transitions from cloud AI to local edge models without interrupting navigation or safety",
        "1. network_sim_node monitors round-trip latency, jitter, and packet loss on communication channel.\n"
        "2. Evaluates connection thresholds:\n"
        "   - ONLINE: Latency < 100ms, loss < 5% -> Full server VLM offloading and fleet coordination enabled.\n"
        "   - PARTIAL: Latency 100-300ms, loss 5-15% -> Non-critical cloud queries disabled; robot switches to local STT/TTS and deterministic planners.\n"
        "   - OFFLINE: Latency > 300ms, loss > 15%, or server disconnected -> Complete edge autonomy mode.\n"
        "3. Broadcasts failover mode on /gracemo/network_state and /gracemo/server_status.\n"
        "4. planner_node and llm_node detect offline status and immediately route all semantic reasoning to local heuristics or offline quantized SLM.\n"
        "5. Safety servo loop continues operating locally at 50 Hz with zero reliance on cloud connectivity."
    ),
    (
        "WF-10",
        "Automated Fault Injection, Self-Diagnosis & Resilience Validation",
        "Operator Injects Fault -> Sensor/Hardware Degradation -> Diagnostics Detection -> Failsafe Mitigation -> Replay Logging",
        "Operator injects LiDAR failure or wheel motor stall during active navigation mission",
        "Robot diagnoses failure, switches to secondary sensor streams, enters cautious safety speed, and logs event for analytics",
        "1. Operator selects fault in Command Center Fault Injection panel (e.g. 'lidar_failure' or 'wheel_fl_failure') and clicks Inject.\n"
        "2. fault_injection_node receives /gracemo/inject_fault and records active fault state in /gracemo/fault_state.\n"
        "3. virtual_space_node simulates the hardware defect:\n"
        "   - If lidar_failure: /scan publishes max range noise; sensor snapshot marks lidar_failed=True.\n"
        "   - If wheel_fl_failure: FL wheel RPM drops to 0; robot drive dynamics introduce steering drag.\n"
        "4. safety_servo_node detects anomalous sensor inputs; reduces max allowable speed limit from 0.95 m/s to 0.25 m/s.\n"
        "5. analytics_node logs diagnostic anomaly with timestamp to /workspace/gracemo_data/logs/.\n"
        "6. When operator clicks 'Clear Fault', system restores sensor pipelines and validates recovery state."
    )
]

def style_header_cell(cell, fill_hex="001F4E78", font_size=11):
    cell.font = Font(name="Calibri", size=font_size, bold=True, color="00FFFFFF")
    cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="00D9D9D9")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_data_cell(cell, is_even=False, align_left=True):
    cell.font = Font(name="Calibri", size=10, bold=False, color="00000000")
    bg_hex = "00F2F2F2" if is_even else "00FFFFFF"
    cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
    cell.alignment = Alignment(horizontal="left" if align_left else "center", vertical="top", wrap_text=True)
    thin = Side(border_style="thin", color="00D9D9D9")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def populate_sheet(wb, title, headers, data, header_fill="001F4E78", col_widths=None):
    if title in wb.sheetnames:
        ws = wb[title]
        wb.remove(ws)
    
    ws = wb.create_sheet(title=title)
    ws.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws.merge_cells("A1:K1" if len(headers) >= 11 else f"A1:{get_column_letter(len(headers))}1")
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"GraceEMO LPU Digital Twin & Autonomous Robotics — {title.replace('_', ' ')}"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="00FFFFFF")
    title_cell.fill = PatternFill(start_color="0016365C", end_color="0016365C", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 36
    
    # Subtitle Banner
    ws.merge_cells("A2:K2" if len(headers) >= 11 else f"A2:{get_column_letter(len(headers))}2")
    sub_cell = ws.cell(row=2, column=1)
    sub_cell.value = "Comprehensive Technical Specifications, Architectural Blueprints, Core Mathematical Algorithms & Operational Pipelines"
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="00D9D9D9")
    sub_cell.fill = PatternFill(start_color="0016365C", end_color="0016365C", fill_type="solid")
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20
    
    # Blank row
    ws.row_dimensions[3].height = 10
    
    # Header Row
    header_row_idx = 4
    ws.row_dimensions[header_row_idx].height = 28
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=header_row_idx, column=col_idx)
        c.value = h
        style_header_cell(c, fill_hex=header_fill, font_size=10)
    
    # Data Rows
    for row_idx, row_data in enumerate(data, start=header_row_idx + 1):
        ws.row_dimensions[row_idx].height = 70 if "Workflows" in title else 60
        is_even = (row_idx % 2 == 0)
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.value = val
            align_l = (col_idx not in [1, 2, 4, 11])
            style_data_cell(c, is_even=is_even, align_left=align_l)
            
    # Set column widths
    if col_widths:
        for col_idx, w in enumerate(col_widths, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = w

def main():
    print(f"Loading {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # 1. Code_Modules_Deep_Dive
    headers_mod = [
        "Module ID",
        "Subsystem / Layer",
        "Source File Path",
        "Tech Stack / Language",
        "Architectural Purpose & Responsibility",
        "Key Classes, Structs & Functions",
        "Inputs, Subscribed Topics & Config",
        "Outputs, Published Topics & Services",
        "Core Algorithm & Mathematical Implementation",
        "Inter-Module Relationships & Failure Fallbacks",
        "Operational Status"
    ]
    widths_mod = [12, 22, 38, 24, 38, 42, 36, 36, 55, 42, 18]
    
    print("Generating Code_Modules_Deep_Dive sheet...")
    populate_sheet(
        wb,
        title="Code_Modules_Deep_Dive",
        headers=headers_mod,
        data=MODULES_DATA,
        header_fill="001F4E78",
        col_widths=widths_mod
    )
    
    # 2. End_to_End_System_Workflows
    headers_wf = [
        "Workflow ID",
        "Pipeline Name",
        "High-Level Dataflow Trajectory",
        "Trigger Event & Preconditions",
        "Expected Physical & Digital Outcome",
        "Step-by-Step Technical Execution Logic & Protocol"
    ]
    widths_wf = [14, 32, 48, 38, 42, 75]
    
    print("Generating End_to_End_System_Workflows sheet...")
    populate_sheet(
        wb,
        title="End_to_End_System_Workflows",
        headers=headers_wf,
        data=WORKFLOWS_DATA,
        header_fill="00375623",
        col_widths=widths_wf
    )
    
    # 3. Update DT_Upgrade_Log with latest deliverables
    if "DT_Upgrade_Log" in wb.sheetnames:
        ws_upg = wb["DT_Upgrade_Log"]
        existing_comps = [ws_upg.cell(row=r, column=1).value for r in range(2, ws_upg.max_row + 1)]
        latest_upgrades = [
            ("InspectorState Telemetry Sync", "MODIFIED", "gracemo_brain/planner_node.py", "✅ Done", "Phase 6", "Unified pose, battery, obstacle, person, voice into cognitive telemetry", "Verified"),
            ("Memory-Driven Named Places", "MODIFIED", "gracemo_memory/memory_node.py", "✅ Done", "Phase 6", "Decoupled destination waypoints from hardcoded strings to SQLite memory recall", "Verified"),
            ("Rust Kernel ROS 2 Bridge", "NEW", "adapters/robot-bridge/gracemo_bridge", "✅ Done", "Phase 6", "Bi-directional bridge connecting ROS 2 topic graph with Rust Tokio EventBus (/emit & SSE)", "Verified"),
            ("Gemini 2.0 Flash Primary Cortex", "MODIFIED", "gracemo_brain/llm_node.py", "✅ Done", "Phase 6", "Primary cognitive cortex running Gemini 2.0 Flash with structured JSON intent schema", "Verified"),
            ("Command Center 4-Wheel Drive HUD", "MODIFIED", "gracemo_gazebo/web/index.html & app.js", "✅ Done", "Phase 6", "4-wheel RPM telemetry HUD, active mission checkpoints (✓/▶/○), and E-STOP halt", "Verified"),
            ("Neck Pan/Tilt & 90° Hand Servo Mux", "MODIFIED", "gracemo_control/safety_servo_node.cpp", "✅ Done", "Phase 6", "50Hz safety supervisor executing neck pan/tilt gaze and 90° hand gestures", "Verified")
        ]
        sample_r = ws_upg.max_row
        for upg in latest_upgrades:
            if upg[0] not in existing_comps:
                ws_upg.append(list(upg))
                new_r = ws_upg.max_row
                for c_idx in range(1, len(upg) + 1):
                    src_c = ws_upg.cell(row=sample_r, column=c_idx)
                    dst_c = ws_upg.cell(row=new_r, column=c_idx)
                    if src_c.has_style:
                        dst_c.font = openpyxl.styles.Font(name=src_c.font.name or "Calibri", size=src_c.font.size or 10)
                        dst_c.fill = openpyxl.styles.PatternFill(start_color=src_c.fill.start_color, end_color=src_c.fill.end_color, fill_type=src_c.fill.fill_type)
                print(f"Added upgrade log entry: {upg[0]}")
    
    print(f"Saving updated workbook to {EXCEL_PATH}...")
    wb.save(EXCEL_PATH)
    print("✅ Master Database successfully enriched with deep-dive technical explanations!")

if __name__ == "__main__":
    main()
